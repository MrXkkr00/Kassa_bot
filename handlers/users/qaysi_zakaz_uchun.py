import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardRemove

from handlers.users.start import AsosiyState
from keyboards.default.keyboards import only_bosh_menu
from loader import dp
from keyboards.default.keyboards import bosh_menu


@dp.message_handler(text=["Taʼminot", "Oylik", "Oshxona", "Avtomobil", "Uskunalarga xarajat", "Marketing", "Malaka oshirish",
                          "Arenda", "Soliq", "Safar xarajatlari", "Dvidend", "Boshqa xarajatlar", "Kassalararo", "Dollar maydalash",
                          "Turli shaxslar", "Mijozdan kirim", "Turli shaxslar", "Investitsiya"],
                    state=AsosiyState.kirim_harajat_turi)
async def bot_2(message: types.Message, state: FSMContext):
    await state.update_data(
        {"q_y_x_turi": message.text}
    )
    wb_z_u = openpyxl.load_workbook('data/Exel/Zakaz_nomlari.xlsx')
    ws_z_u = wb_z_u.active
    values_1 = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
    text = ""
    text2 = ""
    for i in range(ws_z_u.max_row):
        # print(values_1[i])
        if i < 100:
            text = text + f"<b>{(i + 1)}</b>         {values_1[i]}\n"
        else:
            text2 = text2 + f"<b>{(i + 1)}</b>         {values_1[i]}\n"

    text_1 = ["Oylik", "Dvidend", "Kassalararo", "Dollar maydalash", "Turli shaxslar", "Investitsiya"]
    text_2 = ["Oshxona", "Avtomobil", "Uskunalarga xarajat", "Marketing", "Malaka oshirish", "Arenda", "Soliq"]
    msg = message.text
    if msg in text_1:
        await state.update_data(
            {"q_z_u": "Alohida"}
        )
        await message.answer(f"Tafsilotlari", reply_markup=only_bosh_menu)
        await AsosiyState.tafsilotlar.set()

    elif msg in text_2:
        await state.update_data(
            {"q_z_u": "Umumiy"}
        )
        await message.answer(f"Tafsilotlari", reply_markup=only_bosh_menu)
        await AsosiyState.tafsilotlar.set()

    else:
        await message.answer(f"{text}\n{text2}")
        await message.answer(f"Qaysi zakaz uchun", reply_markup=only_bosh_menu)
        await AsosiyState.qaysi_zakaz_uchun.set()


@dp.message_handler(lambda message: not message.text.isdigit(), state=AsosiyState.qaysi_zakaz_uchun)
async def proc_inva123lid(message: types.Message):
    return await message.reply("Bu yerga faqat raqam yuboring :\nNamuna : 5")


@dp.message_handler(lambda message: message.text.isdigit(), state=AsosiyState.qaysi_zakaz_uchun)
async def bot_krim(message: types.Message, state: FSMContext):
    try:
        wb_z_u_n = openpyxl.load_workbook('data/Exel/Zakaz_nomlari.xlsx')
        ws_z_u_n = wb_z_u_n.active
        values_1_n = [ws_z_u_n.cell(row=i, column=1).value for i in range(1, ws_z_u_n.max_row + 1)]
        text = int(message.text)
        q_z_u = values_1_n[text - 1]
    
        await state.update_data(
            {"q_z_u": q_z_u}
        )
        await message.answer(f"Tafsilotlari", reply_markup=only_bosh_menu)
        await AsosiyState.tafsilotlar.set()
    except:
        await message.answer(f"Ma'lumotlar xato kiritildi\nIltimos boshidan boshlang", reply_markup=bosh_menu)
        await state.finish()
