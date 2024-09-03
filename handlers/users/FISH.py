import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardRemove

from handlers.users.start import AsosiyState
from keyboards.default.keyboards import tasdik, only_bosh_menu
from loader import dp


@dp.message_handler(lambda message: message.text.isdigit(), state=AsosiyState.summa)
async def bot_valyuta(message: types.Message, state: FSMContext):
    await state.update_data(
        {"summa": message.text}
    )
    text1 = ["Taʼminot", "Arenda",  "Oshxona", "Uskunalarga xarajat", "Marketing", "Mijozdan kirim", "Investitsiya",
                          "Soliq", "Safar xarajatlari", "Dvidend", "Boshqa xarajatlar", "Dollar maydalash"]
    text2 = "Oylik"
    text3 = "Turli shaxslar"
    text4 = "Kassalararo"
    text5 = "Avtomobil"
    data = await state.get_data()
    q_y_x_turi = str(data.get("q_y_x_turi"))

    if q_y_x_turi in text1:
        await state.update_data(
            {"fish": q_y_x_turi}
        )
        data = await state.get_data()
        q_y_x_turi = str(data.get("q_y_x_turi"))
        val = str(data.get("val"))
        q_z_u = str(data.get("q_z_u"))
        summa = str(data.get("summa"))
        fish = str(data.get("fish"))
        oper = str(data.get("oper"))
        tafsil = str(data.get("tafsil"))
        kassa = str(data.get("kassa"))
        await message.answer(f"Operatisya  :  {oper}\n"
                             f"Qassa  : {kassa}\n"
                             f"Operatsiya turi : {q_y_x_turi}\n"
                             f"Zakaz : {q_z_u}\n"
                             f"Tafsilotlar   : {tafsil}\n"
                             f"Valyuta  :  {val}\n"
                             f"Summa  : {summa}\n"
                             f"FIO  : {fish}\n")
        await message.answer(f"Shu ma'lumotlarni tasdiqlaysizmi", reply_markup=tasdik)
        await AsosiyState.finish.set()
    else:
        if q_y_x_turi == text2:
            wb_oy = openpyxl.load_workbook('data/Exel/FISHxodimlar.xlsx')
            ws_oy = wb_oy.active
            values_1 = [ws_oy.cell(row=i, column=1).value for i in range(1, ws_oy.max_row + 1)]
            text = ""
            for i in range(ws_oy.max_row):
                text = text + f"<b>{(i + 1)}</b>         {values_1[i]}\n"
            await message.answer(text)
            await message.answer(f"Ф.И.О ")

        elif q_y_x_turi == text3:
            wb_t_sh = openpyxl.load_workbook('data/Exel/Turli_shaxslar.xlsx')
            ws_t_sh = wb_t_sh.active
            values_1 = [ws_t_sh.cell(row=i, column=1).value for i in range(1, ws_t_sh.max_row + 1)]
            text3_1 = ""
            for i in range(ws_t_sh.max_row):
                text3_1 = text3_1 + f"<b>{(i + 1)}</b>         {values_1[i]}\n"
            await message.answer(text3_1)
            await message.answer(f"F.I.O")

        elif q_y_x_turi == text4:
            wb_k_o = openpyxl.load_workbook('data/Exel/Kassalararo.xlsx')
            ws_k_o = wb_k_o.active
            values_1 = [ws_k_o.cell(row=i, column=1).value for i in range(1, ws_k_o.max_row + 1)]
            text4_1 = ""
            for i in range(ws_k_o.max_row):
                text4_1 = text4_1 + f"<b>{(i + 1)}</b>         {values_1[i]}\n"
            await message.answer(text4_1)
            await message.answer(f"F.I.O")

        elif q_y_x_turi == text5:
            wb_a = openpyxl.load_workbook('data/Exel/Avtamabillar.xlsx')
            ws_a = wb_a.active
            values_1 = [ws_a.cell(row=i, column=1).value for i in range(1, ws_a.max_row + 1)]
            text5_1 = ""
            for i in range(ws_a.max_row):
                text5_1 = text5_1 + f"<b>{(i + 1)}</b>         {values_1[i]}\n"
            await message.answer(text5_1)
            await message.answer(f"F.I.O")
        else:
            await message.answer(f"F.I.O", reply_markup=only_bosh_menu)
        await AsosiyState.fish.set()


