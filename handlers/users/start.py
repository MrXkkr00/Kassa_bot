from time import strftime, gmtime

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.builtin import CommandStart
from aiogram.dispatcher.filters.state import StatesGroup, State

from data.config import kassirlar
from handlers.users.transliterate import to_cyrillic, to_latin
from keyboards.default.keyboards import bosh_menu, kirim_keyboard, chiqim_keyboard, valyuta, tasdik, only_bosh_menu, \
    xisob_raqamlar, kartalar
from loader import dp, bot
from utils.db_api.xisobot import Database_Xisob


@dp.message_handler(text="/start")
async def bot_sdfsdfart(message: types.Message):
    # print(message.from_user.id)
    await message.answer(f"Assalomu Alaykum!", reply_markup=bosh_menu)


@dp.message_handler(text="start")
async def bot_4656465dfart(message: types.Message):
    # print(message.from_user.id)
    await message.answer(f"Assalomu Alaykum!", reply_markup=bosh_menu)


@dp.message_handler(text="toza")
async def bot_start(message: types.Message):
    # print(message.from_user.id)
    await message.answer(f"Toza", reply_markup=types.ReplyKeyboardRemove())


class AsosiyState(StatesGroup):
    kassa = State()
    kirim_harajat_turi = State()
    qaysi_zakaz_uchun = State()
    tafsilotlar = State()
    valyuta = State()
    summa = State()
    fish = State()
    finish = State()


# buni o'zgartirmang botni ishlamidigan qilib qo'yaman
@dp.message_handler(text="Owner")
async def Asoschi(message: types.Message):
    await message.answer(f"Qadomboyev Rasulbek \n"
                         f"T.number: +998993321038\n"
                         f"T.number: +998334040300\n"
                         f"telegram: @QRasulbek")


@dp.message_handler(text="🏠Bosh menu",
                    state=[AsosiyState.kassa, AsosiyState.kirim_harajat_turi, AsosiyState.qaysi_zakaz_uchun,
                           AsosiyState.tafsilotlar, AsosiyState.valyuta, AsosiyState.summa, AsosiyState.fish,
                           AsosiyState.finish])
async def boshdsadnu(message: types.Message, state: FSMContext):
    await message.answer(f"🏠Bosh menu", reply_markup=bosh_menu)
    await state.finish()


@dp.message_handler(text="Chiqim")
async def bot_krim(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(
        {"oper": message.text}
    )
    try:
        kassir = kassirlar[str(user_id)]

    except:
        kassir = "None"

    if kassir == "Bobomurod Sirojiddinov":
        await message.answer(f"Xisob raqamlari", reply_markup=xisob_raqamlar)
        await AsosiyState.kassa.set()

    elif kassir == "Usmonjon Nematov":
        await message.answer(f"Qarparativ Kartalar", reply_markup=kartalar)
        await AsosiyState.kassa.set()
    else:
        await state.update_data(
            {"kassa": kassir}
        )
        await message.answer(f"Kirim yoki xarajat turlari", reply_markup=chiqim_keyboard)
        await AsosiyState.kirim_harajat_turi.set()


@dp.message_handler(state=AsosiyState.kassa)
async def bot_krim(message: types.Message, state: FSMContext):
    data = await state.get_data()
    oper = str(data.get("oper"))
    await state.update_data(
        {"kassa": message.text}
    )
    if oper == "Kirim":
        await message.answer(f"Kirim yoki xarajat turlari", reply_markup=kirim_keyboard)
        await AsosiyState.kirim_harajat_turi.set()
    else:
        await message.answer(f"Kirim yoki xarajat turlari", reply_markup=chiqim_keyboard)
        await AsosiyState.kirim_harajat_turi.set()


@dp.message_handler(text="Kirim")
async def bot_chiqim(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(
        {"oper": message.text}
    )
    try:
        kassir = kassirlar[str(user_id)]
    except:
        kassir = "None"

    if kassir == "Bobomurod Sirojiddinov":
        await message.answer(f"Хисоб рақамлар", reply_markup=xisob_raqamlar)
        await AsosiyState.kassa.set()

    elif kassir == "Usmonjon Nematov":
        await message.answer(f"Qarparativ Kartalar", reply_markup=kartalar)
        await AsosiyState.kassa.set()
    else:
        await state.update_data(
            {"kassa": kassir}
        )
        await message.answer(f"Kirim yoki xarajat turlari", reply_markup=kirim_keyboard)
        await AsosiyState.kirim_harajat_turi.set()


@dp.message_handler(state=AsosiyState.tafsilotlar)
async def bot_krim(message: types.Message, state: FSMContext):
    await state.update_data(
        {"tafsil": message.text}
    )

    data = await state.get_data()
    kassa = str(data.get("kassa"))

    if kassa in ['"Yasmin Mebel New Group"', '"Moy Gorod Invest"', '"Moy Gorod Invest"',
                 '"Buyuk Sahovatli Zamin"', '"House Yasmina"', '"Saxovat Madad Nur Fayz"',
                 'Karta Yanmin', "Karta Moy Gorod", "Karta Xandamir", "Karta House Yasmin", "Karta Saxovat"]:
        await state.update_data(
            {"val": "sum"}
        )
        await message.answer(f"Сумма", reply_markup=only_bosh_menu)
        await AsosiyState.summa.set()

    else:
        await message.answer(f"Валюта", reply_markup=valyuta)
        await AsosiyState.valyuta.set()


@dp.message_handler(text=["sum", "$"], state=AsosiyState.valyuta)
async def bot_valyuta(message: types.Message, state: FSMContext):
    await state.update_data(
        {"val": message.text}
    )
    await message.answer(f"Summa", reply_markup=only_bosh_menu)
    await AsosiyState.summa.set()


@dp.message_handler(lambda message: not message.text.isdigit(), state=AsosiyState.summa)
async def proc_inva123lid(message: types.Message):
    return await message.reply("Bu yerga faqat raqam yuboring:\nNamuna : 5")


@dp.message_handler(lambda message: message.text.isdigit(), state=AsosiyState.fish)
async def bot_valyuta(message: types.Message, state: FSMContext):
    text2 = "Oylik"
    text3 = "Turli shaxslar"
    text4 = "Kassalararo"
    text5 = "Avtomobil"
    data = await state.get_data()
    q_y_x_turi = str(data.get("q_y_x_turi"))
    text = int(message.text)

    try:
        if q_y_x_turi == text2:
            wb_oy_1 = openpyxl.load_workbook('data/Exel/FISHxodimlar.xlsx')
            ws_oy_1 = wb_oy_1.active
            values_oy_save = [ws_oy_1.cell(row=i, column=1).value for i in range(1, ws_oy_1.max_row + 1)]

            await state.update_data(
                {"fish": values_oy_save[text - 1]}
            )

        elif q_y_x_turi == text3:
            wb_t_sh_1 = openpyxl.load_workbook('data/Exel/Turli_shaxslar.xlsx')
            ws_t_sh_1 = wb_t_sh_1.active
            values_t_sh_save = [ws_t_sh_1.cell(row=i, column=1).value for i in range(1, ws_t_sh_1.max_row + 1)]
            await state.update_data(
                {"fish": values_t_sh_save[text - 1]})

        elif q_y_x_turi == text4:
            wb_k_o_1 = openpyxl.load_workbook('data/Exel/Kassalararo.xlsx')
            ws_k_o_1 = wb_k_o_1.active
            values_k_o_save = [ws_k_o_1.cell(row=i, column=1).value for i in range(1, ws_k_o_1.max_row + 1)]
            await state.update_data(
                {"fish": values_k_o_save[text - 1]})

        elif q_y_x_turi == text5:
            wb_a_1 = openpyxl.load_workbook('data/Exel/Avtamabillar.xlsx')
            ws_a_1 = wb_a_1.active
            values_a_save = [ws_a_1.cell(row=i, column=1).value for i in range(1, ws_a_1.max_row + 1)]
            await state.update_data(
                {"fish": values_a_save[text - 1]})

        data = await state.get_data()
        q_y_x_turi = str(data.get("q_y_x_turi"))
        val = str(data.get("val"))
        q_z_u = str(data.get("q_z_u"))
        summa = str(data.get("summa"))
        fish = str(data.get("fish"))
        oper = str(data.get("oper"))
        tafsil = str(data.get("tafsil"))
        kassa = str(data.get("kassa"))

        await message.answer(f"Operatisya  :  {oper}\n"
                             f"Qassa  : {kassa}\n"
                             f"Operatsiya turi : {q_y_x_turi}\n"
                             f"Zakaz : {q_z_u}\n"
                             f"Tafsilotlar   : {tafsil}\n"
                             f"Valyuta  :  {val}\n"
                             f"Summa  : {summa}\n"
                             f"FIO  : {fish}\n")
        await message.answer(f"Shu ma'lumotlarni tasdiqlaysizmi", reply_markup=tasdik)
        await AsosiyState.finish.set()
    except:
        await message.answer(f"Ma'lumotlar xato kiritildi\nIltimos boshidan boshlang", reply_markup=bosh_menu)
        await state.finish()


@dp.message_handler(state=AsosiyState.fish)
async def bot_valy12ta(message: types.Message, state: FSMContext):
    await state.update_data(
        {"fish": message.text}
    )
    data = await state.get_data()
    q_y_x_turi = str(data.get("q_y_x_turi"))
    val = str(data.get("val"))
    q_z_u = str(data.get("q_z_u"))
    summa = str(data.get("summa"))
    fish = str(data.get("fish"))
    oper = str(data.get("oper"))
    tafsil = str(data.get("tafsil"))
    kassa = str(data.get("kassa"))
    await message.answer(f"Operatisya  :  {oper}\n"
                         f"Qassa  : {kassa}\n"
                         f"Operatsiya turi : {q_y_x_turi}\n"
                         f"Zakaz : {q_z_u}\n"
                         f"Tafsilotlar   : {tafsil}\n"
                         f"Valyuta  :  {val}\n"
                         f"Summa  : {summa}\n"
                         f"FIO  : {fish}\n")
    await message.answer(f"Шу малумотларни тасдиклайсизми", reply_markup=tasdik)
    await AsosiyState.finish.set()


@dp.message_handler(text="tasdiklash", state=AsosiyState.finish)
async def bot_val32yuta(message: types.Message, state: FSMContext):
    data = await state.get_data()
    oper = str(data.get("oper"))
    kassa = str(data.get("kassa"))
    q_y_x_turi = str(data.get("q_y_x_turi"))
    q_z_u = str(data.get("q_z_u"))
    tafsil = str(data.get("tafsil"))
    val = str(data.get("val"))
    summa = str(data.get("summa"))
    fish = str(data.get("fish"))
    soat = (int(strftime('%H', gmtime())) + 5) % 24
    vaqt = str(strftime(f"%d/%m/%Y", gmtime()))

    await bot.send_message(-1001808463960, f"<b>Операция</b>  :  {oper}\n"
                                           f"<b>Касса</b>  :  {kassa}\n"
                                           f"<b>Операция тури</b> :  {q_y_x_turi}\n"
                                           f"<b>Заказ</b> :  {q_z_u}\n"
                                           f"<b>Tafsilotlari</b>   :   {tafsil}\n"
                                           f"<b>Валюта</b>  :  {val}\n"
                                           f"<b>Сумма</b>  :   {summa}\n"
                                           f"<b>ФИО</b>  :  {fish}\n"
                                           f"<b>Time</b> :  {vaqt}\n")

    def Togirlash(msg):
        lenthe = len(msg)
        text = ""
        for i in range(lenthe):
            if msg[i] == "қ":
                text = text + "к"
            elif msg[i] == "Қ":
                text = text + "К"

            elif msg[i] == "Ғ":
                text = text + "Г"
            elif msg[i] == "ғ":
                text = text + "г"

            elif msg[i] == "ҳ":
                text = text + "х"
            elif msg[i] == "Ҳ":
                text = text + "Х"

            elif msg[i] == "Ў":
                text = text + "У"
            elif msg[i] == "ў":
                text = text + "у"

            else:
                text = text + msg[i]
        return text

    # oper = Togirlash(oper)
    # kassa = Togirlash(kassa)
    # q_y_x_turi = Togirlash(q_y_x_turi)
    # q_z_u = Togirlash(q_z_u)
    # tafsil = Togirlash(tafsil)
    # val = Togirlash(val)
    # summa = Togirlash(summa)
    # fish = Togirlash(fish)

    f_oper = open("./data/Text/1_oper.txt", "a", encoding="utf-8")
    f_oper.write(f"\n{oper}")
    f_oper.close()

    f_kassa = open("./data/Text/2_kassa.txt", "a", encoding="utf-8")
    f_kassa.write(f"\n{kassa}")
    f_kassa.close()

    f_3_kyxt = open("./data/Text/3_kyxt.txt", "a", encoding="utf-8")
    f_3_kyxt.write(f"\n{q_y_x_turi}")
    f_3_kyxt.close()

    f_4_qzu = open("./data/Text/4_qzu.txt", "a", encoding="utf-8")
    f_4_qzu.write(f"\n{q_z_u}")
    f_4_qzu.close()

    f_5_tafsil = open("./data/Text/5_tafsil.txt", "a", encoding="utf-8")
    f_5_tafsil.write(f"\n{tafsil}")
    f_5_tafsil.close()

    f_6_val = open("./data/Text/6_val.txt", "a", encoding="utf-8")
    f_6_val.write(f"\n{val}")
    f_6_val.close()

    f_7_summa = open("./data/Text/7_summa.txt", "a", encoding="utf-8")
    f_7_summa.write(f"\n{summa}")
    f_7_summa.close()

    f_8_fish = open("./data/Text/8_fish.txt", "a", encoding="utf-8")
    f_8_fish.write(f"\n{fish}")
    f_8_fish.close()

    f_vaqt = open("./data/Text/9_vaqt.txt", "a", encoding="utf-8")
    f_vaqt.write(f"\n{vaqt}")
    f_vaqt.close()
    await message.answer(f"✅", reply_markup=bosh_menu)

    await state.finish()


@dp.message_handler(text="qayta kiritish", state=AsosiyState.finish)
async def bot_valyuta(message: types.Message, state: FSMContext):
    await message.answer(f"Eski ma'lumotlar o'chirildi\n"
                         f"Yangi ma'lumotlar kiritishingiz mumkin", reply_markup=bosh_menu)
    await state.finish()
