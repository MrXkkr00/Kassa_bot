import openpyxl

wb_z_u = openpyxl.load_workbook('Zakaz_nomlari.xlsx')
ws_z_u = wb_z_u.active
values_1 = [ws_z_u.cell(row=i, column=1).value for i in range(1, ws_z_u.max_row + 1)]
print(ws_z_u.max_row)